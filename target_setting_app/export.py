"""Excel export logic for GCSE and A Level target grades."""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from target_engine import compute_gcse_summary, compute_alevel_summary

AMBER_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BAND_A_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAND_B_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BAND_C_FILL = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
BAND_D_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FOOTER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

_AL_DEPT_FILLS = {
    "A*": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "A":  PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid"),
    "B":  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "C":  PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
    "D":  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "E":  PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
}

_GCSE_DEPT_FILLS = {
    9: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    8: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    7: PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid"),
    6: PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid"),
    5: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    4: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    3: PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
    2: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    1: PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
}


def _is_blank(val) -> bool:
    """Return True if val is NA, None, empty string, or pd.NA."""
    if val is None:
        return True
    try:
        if pd.isna(val):
            return True
    except (TypeError, ValueError):
        pass
    return str(val) == ""


def _auto_fit_columns(ws, min_width: int = 6, max_width: int = 30) -> None:
    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 2, max_width))


def _add_dept_sheet(
    wb,
    title: str,
    students: list[dict],
    grade_fills: dict,
    amber_overrides: bool = False,
    year_str: str = "",
) -> None:
    """Add a per-department sheet to the workbook.

    Parameters
    ----------
    wb:
        The openpyxl Workbook to add the sheet to.
    title:
        Sheet name (truncated to 31 chars to satisfy Excel's limit).
    students:
        List of dicts with keys: surname, forename, grade, and optionally overridden.
    grade_fills:
        Mapping from grade value to PatternFill.
    amber_overrides:
        When True, cells where student["overridden"] is True receive AMBER_FILL.
    year_str:
        Academic year string used in the title cell.
    """
    sheet_title = title[:31]
    ws = wb.create_sheet(sheet_title)

    # Row 1: title merged across cols A-C
    title_cell = ws.cell(row=1, column=1, value=f"Emanuel School — {title} — {year_str}")
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Row 3: headers
    for col_idx, header in enumerate(["Surname", "Forename", "Target"], start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Rows 4+: student data sorted alphabetically
    sorted_students = sorted(students, key=lambda s: (str(s.get("surname", "")), str(s.get("forename", ""))))

    for r_idx, student in enumerate(sorted_students, start=4):
        ws.cell(row=r_idx, column=1, value=str(student.get("surname", "")))
        ws.cell(row=r_idx, column=2, value=str(student.get("forename", "")))

        grade = student.get("grade", "")
        grade_cell = ws.cell(row=r_idx, column=3, value=grade)
        grade_cell.alignment = Alignment(horizontal="center")

        # Determine fill
        if amber_overrides and student.get("overridden", False):
            grade_cell.fill = AMBER_FILL
        else:
            fill = grade_fills.get(grade)
            if fill is not None:
                grade_cell.fill = fill

    _auto_fit_columns(ws)


def export_gcse(
    targets_df: pd.DataFrame,
    overrides: dict | None = None,
    academic_year: str = "",
) -> bytes:
    """Generate Excel for GCSE targets with per-subject department sheets. Returns bytes."""
    overrides = overrides or {}
    subject_cols = [
        c for c in targets_df.columns
        if c not in ("surname", "forename", "form", "overall_score")
    ]
    subject_cols = sorted(subject_cols)
    year_str = academic_year or _current_academic_year()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Matrix View"

    # Title row
    ws1.cell(row=1, column=1, value=f"Emanuel School — GCSE Target Grades {year_str}")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(subject_cols))

    # Header row 3
    header_cols = ["Surname", "Forename", "Form"] + subject_cols
    for col_idx, header in enumerate(header_cols, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    df_sorted = targets_df.sort_values("overall_score", ascending=False, na_position="last")
    data_start_row = 4

    for r_idx, (_, row) in enumerate(df_sorted.iterrows(), start=data_start_row):
        ws1.cell(row=r_idx, column=1, value=str(row.get("surname", "")))
        ws1.cell(row=r_idx, column=2, value=str(row.get("forename", "")))
        ws1.cell(row=r_idx, column=3, value=str(row.get("form", "")))

        for c_idx, subj in enumerate(subject_cols, start=4):
            val = row.get(subj)
            if _is_blank(val):
                continue
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

            # Highlight overrides
            student_key = f"{row['surname']}|{row['forename']}"
            if student_key in overrides and subj in overrides[student_key]:
                cell.fill = AMBER_FILL

    # Summary footer
    footer_start = data_start_row + len(df_sorted) + 2
    summary_df = compute_gcse_summary(targets_df, subject_cols)
    ws1.cell(row=footer_start - 1, column=1, value="Grade Distribution").font = Font(bold=True)

    for r_idx, (_, row) in enumerate(summary_df.iterrows(), start=footer_start):
        ws1.cell(row=r_idx, column=1, value=row["Band"]).font = Font(bold=True)
        ws1.cell(row=r_idx, column=1).fill = FOOTER_FILL
        for c_idx, subj in enumerate(subject_cols, start=4):
            cell = ws1.cell(row=r_idx, column=c_idx, value=row.get(subj, ""))
            cell.fill = FOOTER_FILL
            cell.alignment = Alignment(horizontal="center")
        all_cell = ws1.cell(row=r_idx, column=4 + len(subject_cols), value=row.get("All", ""))
        all_cell.fill = FOOTER_FILL
        all_cell.font = Font(bold=True)
        all_cell.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws1)

    # Sheet 2: Unpivoted
    ws2 = wb.create_sheet("Unpivoted Data")
    ws2.append(["Surname", "Forename", "Form", "Subject", "Target Grade"])
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    unpivot_rows = []
    for _, row in df_sorted.iterrows():
        for subj in subject_cols:
            val = row.get(subj)
            if not _is_blank(val):
                unpivot_rows.append((
                    row.get("surname", ""),
                    row.get("forename", ""),
                    row.get("form", ""),
                    subj,
                    val,
                ))

    unpivot_rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[3])))
    for r in unpivot_rows:
        ws2.append(list(r))

    _auto_fit_columns(ws2)

    # Per-subject department sheets
    for subj in subject_cols:
        dept_students = []
        for _, row in df_sorted.sort_values(["surname", "forename"]).iterrows():
            val = row.get(subj)
            if not _is_blank(val) and str(val) not in ("N/A", ""):
                student_key = f"{row['surname']}|{row['forename']}"
                overridden = (
                    student_key in overrides
                    and subj in overrides.get(student_key, {})
                )
                # Convert grade to int for fill lookup where possible
                grade_key = val
                try:
                    grade_key = int(val)
                except (ValueError, TypeError):
                    pass
                dept_students.append({
                    "surname": str(row.get("surname", "")),
                    "forename": str(row.get("forename", "")),
                    "grade": str(val),
                    "_grade_key": grade_key,
                    "overridden": overridden,
                })
        if dept_students:
            # Build a fills dict keyed by the string grade value for _add_dept_sheet,
            # but also support int-keyed _GCSE_DEPT_FILLS via _grade_key on each student.
            # We pass a string-keyed proxy and handle int lookup inside a wrapper.
            _add_dept_sheet_gcse(
                wb,
                subj,
                dept_students,
                _GCSE_DEPT_FILLS,
                year_str=year_str,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_dept_sheet_gcse(
    wb,
    title: str,
    students: list[dict],
    grade_fills: dict,
    year_str: str = "",
) -> None:
    """Add a per-department GCSE sheet, using int-keyed grade fills."""
    sheet_title = title[:31]
    ws = wb.create_sheet(sheet_title)

    title_cell = ws.cell(row=1, column=1, value=f"Emanuel School — {title} — {year_str}")
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    for col_idx, header in enumerate(["Surname", "Forename", "Target"], start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    sorted_students = sorted(
        students,
        key=lambda s: (str(s.get("surname", "")), str(s.get("forename", ""))),
    )

    for r_idx, student in enumerate(sorted_students, start=4):
        ws.cell(row=r_idx, column=1, value=str(student.get("surname", "")))
        ws.cell(row=r_idx, column=2, value=str(student.get("forename", "")))

        grade = student.get("grade", "")
        grade_cell = ws.cell(row=r_idx, column=3, value=grade)
        grade_cell.alignment = Alignment(horizontal="center")

        if student.get("overridden", False):
            grade_cell.fill = AMBER_FILL
        else:
            grade_key = student.get("_grade_key", grade)
            fill = grade_fills.get(grade_key)
            if fill is not None:
                grade_cell.fill = fill

    _auto_fit_columns(ws)


def export_alevel(
    targets_df: pd.DataFrame,
    overrides: dict | None = None,
    academic_year: str = "",
) -> bytes:
    """Generate Excel for A Level targets with per-subject department sheets. Returns bytes."""
    overrides = overrides or {}
    subject_cols = [
        c for c in targets_df.columns
        if c not in ("surname", "forename", "year_group", "overall_score")
    ]
    subject_cols = sorted(subject_cols)
    year_str = academic_year or _current_academic_year()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Matrix View"

    ws1.cell(row=1, column=1, value=f"Emanuel School — A Level Target Grades {year_str}")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(subject_cols))

    header_cols = ["Surname", "Forename", "Year Group"] + subject_cols
    for col_idx, header in enumerate(header_cols, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    df_sorted = targets_df.sort_values("overall_score", ascending=False, na_position="last")
    data_start_row = 4

    for r_idx, (_, row) in enumerate(df_sorted.iterrows(), start=data_start_row):
        ws1.cell(row=r_idx, column=1, value=str(row.get("surname", "")))
        ws1.cell(row=r_idx, column=2, value=str(row.get("forename", "")))
        ws1.cell(row=r_idx, column=3, value=str(row.get("year_group", "Year 12")))

        for c_idx, subj in enumerate(subject_cols, start=4):
            val = row.get(subj)
            if _is_blank(val):
                continue
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

            student_key = f"{row['surname']}|{row['forename']}"
            if student_key in overrides and subj in overrides[student_key]:
                cell.fill = AMBER_FILL

    footer_start = data_start_row + len(df_sorted) + 2
    summary_df = compute_alevel_summary(targets_df, subject_cols)
    ws1.cell(row=footer_start - 1, column=1, value="Grade Distribution").font = Font(bold=True)

    for r_idx, (_, row) in enumerate(summary_df.iterrows(), start=footer_start):
        ws1.cell(row=r_idx, column=1, value=row["Band"]).font = Font(bold=True)
        ws1.cell(row=r_idx, column=1).fill = FOOTER_FILL
        for c_idx, subj in enumerate(subject_cols, start=4):
            cell = ws1.cell(row=r_idx, column=c_idx, value=row.get(subj, ""))
            cell.fill = FOOTER_FILL
            cell.alignment = Alignment(horizontal="center")
        all_cell = ws1.cell(row=r_idx, column=4 + len(subject_cols), value=row.get("All", ""))
        all_cell.fill = FOOTER_FILL
        all_cell.font = Font(bold=True)
        all_cell.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws1)

    # Sheet 2
    ws2 = wb.create_sheet("Unpivoted Data")
    ws2.append(["Surname", "Forename", "Year Group", "Subject", "Target Grade"])
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    unpivot_rows = []
    for _, row in df_sorted.iterrows():
        for subj in subject_cols:
            val = row.get(subj)
            if not _is_blank(val):
                unpivot_rows.append((
                    row.get("surname", ""),
                    row.get("forename", ""),
                    row.get("year_group", "Year 12"),
                    subj,
                    val,
                ))

    unpivot_rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[3])))
    for r in unpivot_rows:
        ws2.append(list(r))

    _auto_fit_columns(ws2)

    # Per-subject department sheets
    for subj in subject_cols:
        dept_students = []
        for _, row in df_sorted.sort_values(["surname", "forename"]).iterrows():
            val = row.get(subj)
            if not _is_blank(val) and str(val) not in ("N/A", ""):
                student_key = f"{row['surname']}|{row['forename']}"
                dept_students.append({
                    "surname": str(row.get("surname", "")),
                    "forename": str(row.get("forename", "")),
                    "grade": str(val),
                    "overridden": (
                        student_key in overrides
                        and subj in overrides.get(student_key, {})
                    ),
                })
        if dept_students:
            _add_dept_sheet(
                wb,
                subj,
                dept_students,
                _AL_DEPT_FILLS,
                amber_overrides=True,
                year_str=year_str,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _current_academic_year() -> str:
    now = datetime.now()
    if now.month >= 9:
        return f"{now.year}/{str(now.year + 1)[2:]}"
    return f"{now.year - 1}/{str(now.year)[2:]}"
