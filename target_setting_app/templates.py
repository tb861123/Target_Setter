"""Downloadable template files for each upload type."""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
_NOTE_FONT = Font(italic=True, color="888888", size=9)
_TITLE_FONT = Font(bold=True, size=12)


def _hrow(ws, row: int, values: list) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")


def _srow(ws, row: int, values: list) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.fill = _SAMPLE_FILL


def _autofit(ws, mn: int = 8, mx: int = 32) -> None:
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)


def _note(ws, row: int, text: str) -> None:
    c = ws.cell(row, 1, text)
    c.font = _NOTE_FONT


def template_yellis_gcse() -> bytes:
    """
    Yellis GCSE .xlsx template (CEM export format).
    Sheet: Sheet1 — data starts row 4 (rows 1–3 are skipped by the parser).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(1, 1, "Yellis GCSE — Template").font = _TITLE_FONT
    ws.merge_cells("A1:M1")
    _note(ws, 2, "Rows 1–3 are skipped by the parser.  Do not change the sheet name (Sheet1).")
    _note(ws, 3, "Columns must appear in exactly this order starting at column A.")

    headers = [
        "Surname", "Forename", "Form", "Sex",
        "Overall Score", "Overall Band",
        "Vocab Score", "Vocab Band",
        "Maths Score", "Maths Band",
        "Patterns Score", "Patterns Band",
        "Range",
    ]
    _hrow(ws, 3, headers)

    samples = [
        ["Smith",   "Alice",   "10A", "F", 102, "B", 98,  "C", 110, "A", 99,  "B", 15],
        ["Jones",   "Thomas",  "10B", "M", 118, "A", 115, "A", 120, "A", 112, "A", 12],
        ["Okonkwo", "Chidera", "10C", "M",  88, "C",  85, "C",  92, "B",  84, "C", 18],
    ]
    for r, row in enumerate(samples, 4):
        _srow(ws, r, row)

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_yellis_alevel() -> bytes:
    """
    Yellis A Level .xlsx template (CEM export format).
    Sheet: Data — data starts row 4 (rows 1–3 are skipped).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.cell(1, 1, "Yellis A Level — Template").font = _TITLE_FONT
    ws.merge_cells("A1:M1")
    _note(ws, 2, "Rows 1–3 are skipped by the parser.  Do not change the sheet name (Data).")
    _note(ws, 3, "Columns must appear in exactly this order starting at column A.")

    headers = [
        "Code", "Surname", "Firstname", "Gender", "DOB",
        "Overall Score", "Overall Band",
        "Vocab Score", "Vocab Band",
        "Maths Score", "Maths Band",
        "Nonverbal Score", "Nonverbal Band",
    ]
    _hrow(ws, 3, headers)

    samples = [
        ["ES001", "Smith",   "Alice",   "F", "15/09/2007", 128, "A",  125, "A",  132, "A*", 126, "A"],
        ["ES002", "Jones",   "Thomas",  "M", "21/03/2007", 112, "B",  108, "B",  115, "A",  110, "B"],
        ["ES003", "Okonkwo", "Chidera", "M", "05/11/2007",  98, "C",   95, "C",  100, "B",   96, "C"],
    ]
    for r, row in enumerate(samples, 4):
        _srow(ws, r, row)

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_gcse_grades() -> bytes:
    """
    GCSE Grades .xlsx template (iSams import format).
    Sheet: iSams import — header on row 2, data from row 3.
    Long format: one row per student × subject; Surname/Forename blank after first row per student.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "iSams import"

    ws.cell(1, 1, "GCSE Grades Import — iSams Export Format").font = _TITLE_FONT
    ws.merge_cells("A1:D1")
    _note(ws, 1, "")  # overwrite merge note — title is at A1
    ws.cell(1, 1, "GCSE Grades Import — iSams Export Format").font = _TITLE_FONT

    _hrow(ws, 2, ["Surname", "Forename", "Subject", "Grade"])

    data = [
        # Surname/Forename left blank after first row per student (forward-filled by parser)
        ["Smith",   "Alice",   "Mathematics",       8],
        ["",        "",        "English Literature", 7],
        ["",        "",        "Biology",            7],
        ["Jones",   "Thomas",  "Mathematics",        6],
        ["",        "",        "History",            8],
        ["",        "",        "Physics",            5],
        ["Okonkwo", "Chidera", "Mathematics",        9],
        ["",        "",        "Chemistry",          8],
        ["",        "",        "English Language",   7],
    ]
    for r, row in enumerate(data, 3):
        for col, val in enumerate(row, 1):
            c = ws.cell(r, col, val)
            if val != "":
                c.fill = _SAMPLE_FILL

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_subject_list_timetable() -> bytes:
    """
    Subject list .xlsx — timetable format (Subject Name | Surname | PreName, sheet 'All').
    Matches the iSams timetable export structure: headers in row 1, data from row 2.
    Assembly, Registration, LifeEd, Games, and Learning Support rows are shown to illustrate
    they are automatically ignored by the parser.
    """
    rows = [
        ["Mathematics",        "Smith",   "Alice"],
        ["Physics",            "Smith",   "Alice"],
        ["Chemistry",          "Smith",   "Alice"],
        ["Assembly",           "Smith",   "Alice"],
        ["Mathematics",        "Jones",   "Thomas"],
        ["History",            "Jones",   "Thomas"],
        ["English Literature", "Jones",   "Thomas"],
        ["Registration",       "Jones",   "Thomas"],
        ["Biology",            "Okonkwo", "Chidera"],
        ["Chemistry",          "Okonkwo", "Chidera"],
        ["Economics",          "Okonkwo", "Chidera"],
        ["LifeEd",             "Okonkwo", "Chidera"],
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All"

    _hrow(ws, 1, ["Subject Name", "Surname", "PreName"])
    for r, row in enumerate(rows, 2):
        _srow(ws, r, row)

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_subject_list_long() -> bytes:
    """
    Subject list CSV — long format (subjects comma-separated in one column).
    """
    df = pd.DataFrame([
        {"Surname": "Smith",   "Forename": "Alice",   "Subjects": "Mathematics, Physics, Chemistry"},
        {"Surname": "Jones",   "Forename": "Thomas",  "Subjects": "Mathematics, History, English Literature"},
        {"Surname": "Okonkwo", "Forename": "Chidera", "Subjects": "Biology, Chemistry, Economics"},
    ])
    return df.to_csv(index=False).encode()


def template_subject_list_wide() -> bytes:
    """
    Subject list .xlsx — wide format (one column per subject, 1 = taking).
    """
    subjects = [
        "Mathematics", "Further Mathematics", "Biology", "Chemistry", "Physics",
        "English Literature", "History", "Geography", "Economics", "Psychology",
        "Art", "Drama", "Music", "Design Technology", "Computing",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject List"

    _hrow(ws, 1, ["Surname", "Forename"] + subjects)

    samples = [
        {"Surname": "Smith",   "Forename": "Alice",   "Mathematics": 1, "Physics": 1, "Chemistry": 1},
        {"Surname": "Jones",   "Forename": "Thomas",  "Mathematics": 1, "History": 1, "English Literature": 1},
        {"Surname": "Okonkwo", "Forename": "Chidera", "Biology": 1, "Chemistry": 1, "Economics": 1},
    ]
    for r_idx, student in enumerate(samples, 2):
        ws.cell(r_idx, 1, student["Surname"]).fill = _SAMPLE_FILL
        ws.cell(r_idx, 2, student["Forename"]).fill = _SAMPLE_FILL
        for c_idx, subj in enumerate(subjects, 3):
            val = student.get(subj, "")
            c = ws.cell(r_idx, c_idx, val)
            c.alignment = Alignment(horizontal="center")
            if val:
                c.fill = _SAMPLE_FILL

    _autofit(ws, mn=6, mx=22)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
